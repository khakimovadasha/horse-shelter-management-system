from __future__ import annotations

from collections.abc import Iterable
from datetime import date, datetime
from decimal import Decimal
from enum import Enum
from io import BytesIO
from pathlib import Path
from typing import Any
from xml.sax.saxutils import escape

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from pydantic import BaseModel
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


KEY_LABELS = {
    "report_type": "Тип отчёта",
    "report_title": "Название отчёта",
    "generated_at": "Дата формирования",
    "generated_by": "Пользователь",
    "filters": "Выбранные фильтры",
    "summary": "Краткая сводка",
    "items": "Данные",
    "item": "Данные",
    "horse": "Лошадь",
    "medical_records": "Медицинские записи",
    "procedures": "Процедуры",
    "tasks": "Задачи",
    "horses_section": "Раздел: лошади",
    "medicine_section": "Раздел: медицина",
    "tasks_section": "Раздел: задачи",
    "finance_section": "Раздел: финансы",
    "users_section": "Раздел: пользователи",
    "status": "Статус",
    "curator_id": "Куратор",
    "curator_name": "Куратор",
    "include_medical_records": "Включить медзаписи",
    "include_procedures": "Включить процедуры",
    "include_tasks": "Включить задачи",
    "include_horses": "Включать лошадей",
    "include_medicine": "Включать медицину",
    "include_finance": "Включать финансы",
    "include_users": "Включать пользователей",
    "medical_records_count": "Количество медзаписей",
    "procedures_count": "Количество процедур",
    "tasks_count": "Количество задач",
    "horse_id": "Лошадь",
    "exact_date": "Точная дата",
    "date_from": "Дата начала",
    "date_to": "Дата окончания",
    "planned_count": "Запланировано",
    "completed_count": "Выполнено",
    "overdue_count": "Просрочено",
    "waiting_count": "Ожидает",
    "in_progress_count": "В работе",
    "total_tasks": "Всего задач",
    "total_items": "Количество записей",
    "title": "Название",
    "description": "Описание",
    "executor_name": "Ответственный",
    "executor_id": "Ответственный",
    "due_date": "Срок выполнения",
    "completed_at": "Дата завершения",
    "procedure_name": "Процедура",
    "planned_date": "Плановая дата",
    "completed_date": "Дата выполнения",
    "procedure_date": "Дата процедуры",
    "add_to_medical_record": "Добавить в медкарту",
    "notes": "Примечание",
    "record_date": "Дата записи",
    "record_type": "Тип записи",
    "name": "Кличка",
    "gender": "Пол",
    "age": "Возраст",
    "breed": "Порода",
    "color": "Масть",
    "arrival_date": "Дата поступления",
    "history": "История",
    "healthy_count": "Здоровы",
    "sick_count": "Болеют",
    "rehabilitation_count": "На реабилитации",
    "deceased_count": "Выбыли",
    "user_id": "Пользователь",
    "user_name": "Пользователь",
    "users": "Пользователи",
    "full_name": "ФИО",
    "role": "Роль",
    "curated_horses_count": "Лошадей",
    "in_progress_tasks_count": "В работе",
    "completed_tasks_count": "Выполнено",
    "operation_type": "Тип операции",
    "category": "Категория",
    "amount": "Сумма",
    "date": "Дата",
    "horse_name": "Лошадь",
    "total_income": "Доходы",
    "total_expense": "Расходы",
    "balance": "Баланс",
    "operations_count": "Количество операций",
    "is_active": "Активен",
    "active_users_count": "Активные пользователи",
    "total_horses": "Всего лошадей",
    "sick_horses": "Больные лошади",
    "procedures_count": "Количество процедур",
    "active_tasks_count": "Активные задачи",
    "overdue_tasks_count": "Просроченные задачи",
    "income_total": "Доходы",
    "expense_total": "Расходы",
}

VALUE_LABELS = {
    "all": "Все",
    "healthy": "Здоров",
    "sick": "Болен",
    "rehabilitation": "Реабилитация",
    "deceased": "Выбыл",
    "male": "Жеребец",
    "female": "Кобыла",
    "planned": "Запланировано",
    "completed": "Выполнено",
    "overdue": "Просрочено",
    "waiting": "Ожидает",
    "in_progress": "В работе",
    "admin": "Администратор",
    "veterinarian": "Ветеринар",
    "user": "Волонтёр",
    "income": "Доход",
    "expense": "Расход",
    "donations": "Пожертвования",
    "sponsorship": "Спонсорство",
    "grants": "Гранты",
    "sales": "Продажи",
    "other_income": "Прочие доходы",
    "medications": "Медикаменты",
    "feed": "Корма",
    "care": "Уход",
    "equipment": "Снаряжение",
    "transport": "Транспорт",
    "repair": "Ремонт",
    "utilities": "Коммунальные расходы",
    "other_expense": "Прочие расходы",
    "inspection": "Осмотр",
    "diagnosis": "Диагноз",
    "treatment": "Лечение",
    "analysis": "Анализ",
    "procedure": "Процедура",
    "note": "Заметка",
}

PDF_HIDDEN_KEYS = {"id", "horse_id", "user_id", "executor_id", "curator_id"}

PDF_FONT_NAME = "AppPdfFont"
PDF_FONT_BOLD_NAME = "AppPdfFontBold"


def ensure_pdf_fonts_registered() -> tuple[str, str]:
    registered_fonts = set(pdfmetrics.getRegisteredFontNames())
    if PDF_FONT_NAME in registered_fonts and PDF_FONT_BOLD_NAME in registered_fonts:
        return PDF_FONT_NAME, PDF_FONT_BOLD_NAME

    font_candidates = [
        (Path("C:/Windows/Fonts/arial.ttf"), Path("C:/Windows/Fonts/arialbd.ttf")),
        (Path("C:/Windows/Fonts/tahoma.ttf"), Path("C:/Windows/Fonts/tahomabd.ttf")),
        (Path("C:/Windows/Fonts/verdana.ttf"), Path("C:/Windows/Fonts/verdanab.ttf")),
    ]

    for regular_path, bold_path in font_candidates:
        if not regular_path.exists() or not bold_path.exists():
            continue

        pdfmetrics.registerFont(TTFont(PDF_FONT_NAME, str(regular_path)))
        pdfmetrics.registerFont(TTFont(PDF_FONT_BOLD_NAME, str(bold_path)))
        return PDF_FONT_NAME, PDF_FONT_BOLD_NAME

    return "Helvetica", "Helvetica-Bold"


def export_report(report: BaseModel, export_format: str) -> tuple[bytes, str, str]:
    report_payload = report.model_dump()
    report_type = report_payload["report_type"]
    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")

    if export_format == "xlsx":
        return (
            build_excel_bytes(report_payload),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            f"{report_type}_{timestamp}.xlsx",
        )

    if export_format == "pdf":
        return (
            build_pdf_bytes(report_payload),
            "application/pdf",
            f"{report_type}_{timestamp}.pdf",
        )

    raise ValueError("Неподдерживаемый формат экспорта")


def build_excel_bytes(report_payload: dict[str, Any]) -> bytes:
    workbook = Workbook()
    overview = workbook.active
    overview.title = "Отчёт"

    row = 1
    row = write_excel_title(overview, row, report_payload["report_title"])
    row = write_excel_metadata(overview, row, report_payload)

    body_keys = [key for key in report_payload.keys() if key not in {"report_type", "report_title", "generated_at", "generated_by", "filters", "summary"}]
    for key in body_keys:
        if report_payload[key] is None:
            continue
        row = render_excel_overview_section(workbook, overview, row, [key], report_payload[key])

    autofit_sheet(overview)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def write_excel_title(sheet, row: int, title: str) -> int:
    sheet.cell(row=row, column=1, value=title)
    sheet.cell(row=row, column=1).font = Font(size=16, bold=True)
    return row + 2


def write_excel_metadata(sheet, row: int, report_payload: dict[str, Any]) -> int:
    metadata = {
        "generated_at": report_payload["generated_at"],
        "generated_by": report_payload["generated_by"],
    }
    row = write_excel_kv_block(sheet, row, "Параметры отчёта", metadata)
    row = write_excel_kv_block(sheet, row, "Фильтры", report_payload["filters"])
    row = write_excel_kv_block(sheet, row, "Краткая сводка", report_payload["summary"])
    return row


def render_excel_overview_section(workbook: Workbook, overview, row: int, path: list[str], value: Any) -> int:
    title = " / ".join(label_key(part) for part in path)

    if is_scalar(value):
        overview.cell(row=row, column=1, value=title)
        overview.cell(row=row, column=2, value=format_value(value))
        return row + 1

    if isinstance(value, dict):
        scalar_part = {
            key: current_value
            for key, current_value in value.items()
            if is_scalar(current_value) and key not in PDF_HIDDEN_KEYS and current_value is not None
        }
        if scalar_part:
            row = write_excel_kv_block(overview, row, title, scalar_part)
        for key, current_value in value.items():
            if is_scalar(current_value):
                continue
            row = render_excel_overview_section(workbook, overview, row, path + [key], current_value)
        return row

    if isinstance(value, list):
        write_excel_table_sheet(workbook, title, value)
        return row

    overview.cell(row=row, column=1, value=title)
    overview.cell(row=row, column=2, value=format_value(value))
    return row + 1


def write_excel_kv_block(sheet, row: int, title: str, data: dict[str, Any]) -> int:
    sheet.cell(row=row, column=1, value=title)
    sheet.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    for key, value in data.items():
        if key in PDF_HIDDEN_KEYS or value is None:
            continue
        key_cell = sheet.cell(row=row, column=1, value=label_key(key))
        value_cell = sheet.cell(row=row, column=2, value=format_value(value))
        key_cell.alignment = Alignment(vertical="top", wrap_text=True)
        value_cell.alignment = Alignment(vertical="top", wrap_text=True)
        row += 1
    return row + 1


def write_excel_table_sheet(workbook: Workbook, title: str, rows: list[Any]) -> None:
    sheet = workbook.create_sheet(title=safe_sheet_name(workbook, title))
    sheet.cell(row=1, column=1, value=title)
    sheet.cell(row=1, column=1).font = Font(size=14, bold=True)

    normalized_rows = [flatten_row(row) for row in rows]
    if not normalized_rows:
        sheet.cell(row=3, column=1, value="Нет данных")
        autofit_sheet(sheet)
        return

    headers = list(dict.fromkeys(key for row in normalized_rows for key in row.keys()))
    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=3, column=column_index, value=label_key(header))
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    for row_index, row_data in enumerate(normalized_rows, start=4):
        for column_index, header in enumerate(headers, start=1):
            cell = sheet.cell(row=row_index, column=column_index, value=row_data.get(header, "—"))
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    autofit_sheet(sheet)


def autofit_sheet(sheet) -> None:
    for column_cells in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            if cell.value is None:
                continue
            value_length = len(str(cell.value))
            if value_length > max_length:
                max_length = value_length
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 14), 50)


def build_pdf_bytes(report_payload: dict[str, Any]) -> bytes:
    buffer = BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=16 * mm,
        leftMargin=16 * mm,
        topMargin=16 * mm,
        bottomMargin=16 * mm,
    )

    regular_font_name, bold_font_name = ensure_pdf_fonts_registered()
    styles = getSampleStyleSheet()
    title_style = styles["Heading1"]
    section_style = styles["Heading2"]
    subsection_style = ParagraphStyle("SubHeading", parent=styles["Heading3"], spaceBefore=8, spaceAfter=4)
    body_style = styles["BodyText"]
    table_header_style = ParagraphStyle(
        "PdfTableHeader",
        parent=body_style,
        fontName=bold_font_name,
    )

    title_style.fontName = bold_font_name
    section_style.fontName = bold_font_name
    subsection_style.fontName = bold_font_name
    body_style.fontName = regular_font_name

    story: list[Any] = [
        Paragraph(report_payload["report_title"], title_style),
        Spacer(1, 6),
    ]

    story.extend(render_pdf_kv_section(
        "Параметры отчёта",
        {
            "generated_at": report_payload["generated_at"],
            "generated_by": report_payload["generated_by"],
        },
        body_style,
        table_header_style,
        section_style,
    ))
    story.extend(render_pdf_kv_section("Фильтры", report_payload["filters"], body_style, table_header_style, section_style))
    story.extend(render_pdf_kv_section("Краткая сводка", report_payload["summary"], body_style, table_header_style, section_style))

    body_keys = [key for key in report_payload.keys() if key not in {"report_type", "report_title", "generated_at", "generated_by", "filters", "summary"}]
    for key in body_keys:
        if report_payload[key] is None:
            continue
        story.extend(
            render_pdf_value(
                [key],
                report_payload[key],
                body_style,
                table_header_style,
                section_style,
                subsection_style,
            )
        )

    document.build(story)
    return buffer.getvalue()


def render_pdf_kv_section(title: str, data: dict[str, Any], body_style, table_header_style, section_style) -> list[Any]:
    elements: list[Any] = [Paragraph(title, section_style), Spacer(1, 3)]
    rows = [[Paragraph("Параметр", table_header_style), Paragraph("Значение", table_header_style)]]
    for key, value in data.items():
        if key in PDF_HIDDEN_KEYS or value is None:
            continue
        rows.append([Paragraph(escape(label_key(key)), body_style), Paragraph(format_pdf_value(value), body_style)])
    elements.append(build_pdf_table(rows))
    elements.append(Spacer(1, 8))
    return elements


def render_pdf_value(path: list[str], value: Any, body_style, table_header_style, section_style, subsection_style) -> list[Any]:
    title = " / ".join(label_key(part) for part in path)
    elements: list[Any] = []

    if is_scalar(value):
        elements.append(Paragraph(title, section_style))
        elements.append(Paragraph(format_value(value), body_style))
        elements.append(Spacer(1, 6))
        return elements

    if isinstance(value, dict):
        elements.extend(
            render_pdf_kv_section(
                title,
                {k: v for k, v in value.items() if is_scalar(v) and k not in PDF_HIDDEN_KEYS},
                body_style,
                table_header_style,
                section_style,
            )
        )
        for key, current_value in value.items():
            if is_scalar(current_value):
                continue
            elements.extend(
                render_pdf_value(
                    path + [key],
                    current_value,
                    body_style,
                    table_header_style,
                    subsection_style,
                    subsection_style,
                )
            )
        return elements

    if isinstance(value, list):
        elements.append(Paragraph(title, section_style))
        normalized_rows = [flatten_pdf_row(item) for item in value]
        if not normalized_rows:
            elements.append(Paragraph("Нет данных", body_style))
            elements.append(Spacer(1, 6))
            return elements

        headers = list(dict.fromkeys(key for row in normalized_rows for key in row.keys()))
        if len(headers) > 6:
            elements.extend(render_pdf_record_blocks(value, body_style, table_header_style, subsection_style))
            return elements

        rows = [[Paragraph(escape(label_key(header)), table_header_style) for header in headers]]
        for row in normalized_rows:
            rows.append([Paragraph(row.get(header, "—"), body_style) for header in headers])
        elements.append(build_pdf_table(rows))
        elements.append(Spacer(1, 8))
        return elements

    elements.append(Paragraph(title, section_style))
    elements.append(Paragraph(format_value(value), body_style))
    elements.append(Spacer(1, 6))
    return elements


def render_pdf_record_blocks(records: list[Any], body_style, table_header_style, subsection_style) -> list[Any]:
    elements: list[Any] = []

    for index, record in enumerate(records, start=1):
        title = resolve_record_title(record, index)
        elements.append(Paragraph(title, table_header_style))

        row = flatten_pdf_row_without_lists(record)
        table_rows = [[Paragraph("Параметр", table_header_style), Paragraph("Значение", table_header_style)]]
        for key, value in row.items():
            table_rows.append([Paragraph(escape(label_key(key)), body_style), Paragraph(value or "—", body_style)])
        elements.append(build_pdf_table(table_rows))

        if isinstance(record, dict):
            for key, value in record.items():
                if not isinstance(value, list) or not value:
                    continue

                elements.append(Paragraph(escape(label_key(key)), subsection_style))
                elements.append(build_pdf_single_column_table(value, body_style, table_header_style))

        elements.append(Spacer(1, 8))

    return elements


def build_pdf_table(rows: list[list[Any]]) -> Table:
    table = Table(rows, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E9EFE7")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D6DEEC")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    return table


def build_pdf_single_column_table(values: list[Any], body_style, table_header_style) -> Table:
    rows: list[list[Any]] = [[Paragraph("Запись", table_header_style)]]

    for value in values:
        if isinstance(value, dict):
            rows.append([Paragraph(format_pdf_record(value), body_style)])
        else:
            rows.append([Paragraph(format_pdf_value(value), body_style)])

    return build_pdf_table(rows)


def resolve_record_title(record: Any, index: int) -> str:
    if isinstance(record, dict):
        horse = record.get("horse")
        if isinstance(horse, dict) and horse.get("name"):
            return escape(str(horse["name"]))

        if record.get("name"):
            return escape(str(record["name"]))

        if record.get("title"):
            return escape(str(record["title"]))

        if record.get("full_name"):
            return escape(str(record["full_name"]))

    return f"Запись {index}"


def flatten_pdf_row_without_lists(value: Any, prefix: str = "") -> dict[str, str]:
    result: dict[str, str] = {}

    if isinstance(value, dict):
        for key, current_value in value.items():
            current_prefix = f"{prefix}.{key}" if prefix else key
            if isinstance(current_value, list):
                continue
            if is_scalar(current_value):
                if key in PDF_HIDDEN_KEYS:
                    continue
                result[current_prefix] = format_pdf_value(current_value)
            elif isinstance(current_value, dict):
                nested_result = flatten_pdf_row_without_lists(current_value, current_prefix)
                if nested_result:
                    result.update(nested_result)
            else:
                result[current_prefix] = format_pdf_value(current_value)
        return result

    result[prefix or "value"] = format_pdf_value(value)
    return result


def flatten_pdf_row(value: Any, prefix: str = "") -> dict[str, str]:
    result: dict[str, str] = {}

    if isinstance(value, dict):
        for key, current_value in value.items():
            current_prefix = f"{prefix}.{key}" if prefix else key
            if is_scalar(current_value):
                if key in PDF_HIDDEN_KEYS:
                    continue
                result[current_prefix] = format_pdf_value(current_value)
            elif isinstance(current_value, dict):
                nested_result = flatten_pdf_row(current_value, current_prefix)
                if nested_result:
                    result.update(nested_result)
            elif isinstance(current_value, list):
                result[current_prefix] = format_pdf_list_value(current_value)
            else:
                result[current_prefix] = format_pdf_value(current_value)
        return result

    result[prefix or "value"] = format_pdf_value(value)
    return result


def format_pdf_record(record: dict[str, Any]) -> str:
    if "full_name" in record and record.get("full_name"):
        return escape(str(record["full_name"]))

    if "first_name" in record and "last_name" in record:
        full_name = f"{record.get('first_name', '')} {record.get('last_name', '')}".strip()
        if full_name:
            return escape(full_name)

    parts: list[str] = []
    for key, value in record.items():
        if key in PDF_HIDDEN_KEYS or value in (None, "", []):
            continue

        if key in {"first_name", "last_name", "full_name"}:
            continue

        parts.append(f"{escape(label_key(key))}: {format_pdf_value(value)}")

    return "<br/>".join(parts) if parts else "—"


def format_pdf_list_value(values: list[Any]) -> str:
    formatted_items: list[str] = []

    for value in values:
        if is_scalar(value):
            formatted_items.append(format_pdf_value(value))
        elif isinstance(value, dict):
            formatted_items.append(format_pdf_record(value))
        else:
            formatted_items.append(format_pdf_value(value))

    return "<br/><br/>".join(item for item in formatted_items if item) if formatted_items else "—"


def format_pdf_value(value: Any) -> str:
    if isinstance(value, dict):
        return format_pdf_record(value)
    if isinstance(value, list):
        return format_pdf_list_value(value)

    return escape(format_value(value))


def safe_sheet_name(workbook: Workbook, desired_name: str) -> str:
    cleaned = desired_name.replace("/", " ").replace("\\", " ").replace(":", " ").strip()
    cleaned = cleaned[:28] or "Sheet"
    name = cleaned
    counter = 1
    while name in workbook.sheetnames:
        suffix = f" {counter}"
        name = f"{cleaned[: 28 - len(suffix)]}{suffix}"
        counter += 1
    return name


def flatten_row(value: Any, prefix: str = "") -> dict[str, str]:
    result: dict[str, str] = {}

    if isinstance(value, dict):
        for key, current_value in value.items():
            current_prefix = f"{prefix}.{key}" if prefix else key
            if is_scalar(current_value):
                if key in PDF_HIDDEN_KEYS:
                    continue
                result[current_prefix] = format_value(current_value)
            elif isinstance(current_value, dict):
                result.update(flatten_row(current_value, current_prefix))
            elif isinstance(current_value, list):
                result[current_prefix] = format_list_value(current_value)
            else:
                result[current_prefix] = format_value(current_value)
        return result

    result[prefix or "value"] = format_value(value)
    return result


def format_list_value(values: list[Any]) -> str:
    formatted_items: list[str] = []
    for value in values:
        if is_scalar(value):
            formatted_items.append(format_value(value))
        elif isinstance(value, dict):
            item_parts = [
                f"{label_key(key)}: {format_value(current_value)}"
                for key, current_value in flatten_row(value).items()
            ]
            formatted_items.append("\n".join(item_parts))
        else:
            formatted_items.append(format_value(value))
    return "\n\n".join(formatted_items) if formatted_items else "—"


def is_scalar(value: Any) -> bool:
    return value is None or isinstance(value, (str, int, float, bool, Decimal, date, datetime, Enum))


def label_key(key: str) -> str:
    if key in KEY_LABELS:
        return KEY_LABELS[key]
    normalized = key.split(".")[-1]
    if normalized in KEY_LABELS:
        return KEY_LABELS[normalized]
    return normalized.replace("_", " ").capitalize()


def format_value(value: Any) -> str:
    if value is None:
        return "—"
    if isinstance(value, bool):
        return "Да" if value else "Нет"
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y %H:%M")
    if isinstance(value, date):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, Decimal):
        return format_decimal(value)
    if isinstance(value, Enum):
        return VALUE_LABELS.get(value.value, value.value)
    if isinstance(value, dict):
        if "full_name" in value and value.get("full_name"):
            return str(value["full_name"])
        if "first_name" in value and "last_name" in value:
            full_name = f"{value.get('first_name', '')} {value.get('last_name', '')}".strip()
            if full_name:
                return full_name
        return "; ".join(f"{label_key(key)}: {format_value(current_value)}" for key, current_value in value.items())
    if isinstance(value, list):
        return format_list_value(value)
    if isinstance(value, str):
        return VALUE_LABELS.get(value, value)
    return str(value)


def format_decimal(value: Decimal) -> str:
    normalized = f"{value:,.2f}".replace(",", " ").replace(".", ",")
    if normalized.endswith(",00"):
        return normalized[:-3]
    return normalized
